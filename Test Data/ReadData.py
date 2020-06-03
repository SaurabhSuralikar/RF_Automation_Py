import openpyxl
wk=openpyxl.load_workbook("C://Users//hp//PycharmProjects//RF_Automation//Test Data//TestData.xlsx")
#sh=wk["Sheet1"]
#print(sh.max_row)
#cell=sh.cell(1,1)
#print(cell.value)

def fetch_no_of_rows(Sheetname):
    sh=wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname,row,cell):
    sh = wk[Sheetname]
    cell=sh.cell(int(row),int(cell))
    return cell.value

print(fetch_no_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1",2,1))
